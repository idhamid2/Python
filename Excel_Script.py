# Script Name: Excel_Script.py
# Description: Script used to compare some fields value and if it is true then copy 
#              value from workbokk-2 to workbook-2 for row respectively.
# Things to change: Chenge the excel file name and workbook-names
# Date:        20-04-2023

from openpyxl import load_workbook

# Load the workbook
workbook = load_workbook(filename='input_file.xlsx')

# Load the worksheets
worksheet1 = workbook['Workbook-1']
worksheet2 = workbook['Workbook-2']

# Iterate over the rows in worksheet1
for row1 in worksheet1.iter_rows():
    for row2 in worksheet2.iter_rows():
        if row1[2].value == row2[1].value and row1[1].value == row2[0].value:
            row1[3].value = row2[4].value
            break

workbook.save(filename='output_file.xlsx')
